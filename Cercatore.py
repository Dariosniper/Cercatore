# Cercatore.py (con miglioramenti strutturali e tracciamento file analizzati)

import os, re, pytesseract, tempfile, subprocess, shutil
import matplotlib.pyplot as plt
from PIL import Image
from PyPDF2 import PdfReader
from docx import Document
from openpyxl import load_workbook
import pandas as pd
from datetime import datetime
from fpdf import FPDF
from tqdm import tqdm
from collections import Counter

# Configurazioni
SUPPORTED_EXTENSIONS = [".txt", ".pdf", ".png", ".jpg", ".jpeg", ".docx", ".xlsx"]
SENSITIVE_PATTERNS = {
    "Codice Fiscale": re.compile(r"[A-Z]{6}[0-9]{2}[A-Z][0-9]{2}[A-Z][0-9]{3}[A-Z]"),
    "Email": re.compile(r"[\w\.-]+@[\w\.-]+"),
    "Telefono": re.compile(r"\b\d{9,11}\b"),
    "IBAN": re.compile(r"\b[A-Z]{2}\d{2}[A-Z0-9]{11,30}\b"),
    "Carta di credito": re.compile(r"\b(?:\d[ -]*?){13,16}\b"),
    "Termini sanitari": re.compile(r"(?i)\b(patologia|farmaco|diagnosi|terapia|esame|tumore|diabete|asma|invalidità)\b"),
    "Religione": re.compile(r"(?i)\b(cristiano|musulmano|ebraico|ateo|cattolico)\b"),
    "Politica": re.compile(r"(?i)\b(partito|comunista|fascista|democratico|lega|movimento 5 stelle)\b"),
    "Sindacati": re.compile(r"(?i)\b(cgil|cisl|uil|sindacato)\b"),
    "Giudiziario": re.compile(r"(?i)\b(condanna|tribunale|avvocato|processo|penale|ricorso)\b"),
    "Password": re.compile(r"(?i)\b(pass|password|pwd|passcode|credenziali|login|cred|accesso|accessi)\b")
}
CLASSIFICATION_LEVELS = {
    "Alto": ["Codice Fiscale", "IBAN", "Carta di credito", "Termini sanitari", "Giudiziario", "Password"],
    "Medio": ["Email", "Telefono", "Religione", "Politica", "Sindacati"]
}
EXCLUDE_FOLDERS = ["$Recycle.Bin", "Program Files", "Windows", "System Volume Information"]


def extract_text_from_file(file_path):
    ext = os.path.splitext(file_path)[1].lower()
    text = ""
    try:
        if ext == ".txt":
            with open(file_path, "r", encoding="utf-8", errors="ignore") as f:
                text = f.read()
        elif ext == ".pdf":
            reader = PdfReader(file_path)
            for page in reader.pages:
                text += page.extract_text() or ""
        elif ext in [".png", ".jpg", ".jpeg"]:
            text = pytesseract.image_to_string(Image.open(file_path))
        elif ext == ".docx":
            doc = Document(file_path)
            text = "\n".join(p.text for p in doc.paragraphs)
        elif ext == ".xlsx":
            wb = load_workbook(file_path, data_only=True)
            for sheet in wb.worksheets:
                for row in sheet.iter_rows(values_only=True):
                    text += " ".join(str(cell) for cell in row if cell) + "\n"
    except (FileNotFoundError, PermissionError, OSError) as e:
        print(f"Errore lettura {file_path}: {e}")
    return text

def classify_risk(found_labels):
    if any(lbl in CLASSIFICATION_LEVELS["Alto"] for lbl in found_labels):
        return "Alto"
    if any(lbl in CLASSIFICATION_LEVELS["Medio"] for lbl in found_labels):
        return "Medio"
    return "Basso"

def scan_directory(base_path, max_size, search_pattern=None):
    results = []
    all_scanned = []

    for root, _, files in os.walk(base_path):
        if any(excl in root for excl in EXCLUDE_FOLDERS):
            continue
        for file in files:
            full_path = os.path.join(root, file)
            ext = os.path.splitext(full_path)[1].lower()
            try:
                size = os.path.getsize(full_path)
            except PermissionError:
                continue
            if ext in SUPPORTED_EXTENSIONS and size <= max_size:
                all_scanned.append(full_path)

    for full in tqdm(all_scanned, desc="Analisi file", unit="file"):
        content = extract_text_from_file(full)
        if search_pattern and not re.search(search_pattern, content, re.IGNORECASE):
            continue
        labels = [label for label, pat in SENSITIVE_PATTERNS.items() if pat.search(content)]
        if labels:
            risk = classify_risk(labels)
            results.append({
                "File": full,
                "Tags": ", ".join(labels),
                "Dimensione": os.path.getsize(full),
                "Rischio": risk
            })
        else:
            results.append({
                "File": full,
                "Tags": "",
                "Dimensione": os.path.getsize(full),
                "Rischio": "Non Classificato"
            })
    return results

def export_results(results, base_output_dir):
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_dir = os.path.join(base_output_dir, f"report_{timestamp}")
    os.makedirs(output_dir, exist_ok=True)

       # Genera grafici
    if results:
        try:
            tag_counter = Counter()
            risk_counter = Counter()
            for row in results:
                tags = [t.strip() for t in re.split(r"[;,]", row.get("Tags", "")) if t.strip()]
                for tag in tags:
                    tag_counter[tag] += 1
                risk = row.get("Rischio", "Non classificato")
                risk_counter[risk] += 1

            # Grafico a torta per rischio
            if risk_counter:
                plt.figure(figsize=(6,6))
                plt.pie(risk_counter.values(), labels=risk_counter.keys(), autopct='%1.1f%%')
                plt.title("Distribuzione per Rischio")
                plt.tight_layout()
                plt.savefig(os.path.join(output_dir, "rischi.png"), bbox_inches='tight', dpi=200)
                plt.close()

            # Grafico a barre per tag
            if tag_counter:
                plt.figure(figsize=(10,6))
                plt.bar(tag_counter.keys(), tag_counter.values())
                plt.xticks(rotation=45, ha='right')
                plt.title("Frequenza dei Dati Sensibili")
                plt.tight_layout()
                plt.savefig(os.path.join(output_dir, "sensibili.png"), bbox_inches='tight', dpi=200)
                plt.close()

        except Exception as e:
            print("Errore generazione grafici:", e)


    excel_path = os.path.join(output_dir, "report.xlsx")
    pdf_path = os.path.join(output_dir, "report.pdf")

    df = pd.DataFrame(results)
    df.to_excel(excel_path, index=False)

    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Arial", size=12)
    pdf.cell(200, 10, txt="Cercatore - Report Sintetico", ln=True, align="C")
    pdf.ln(10)

    for row in results[:30]:
        pdf.multi_cell(0, 10, f"File: {row['File']}\nDati: {row['Tags']}\nRischio: {row['Rischio']}\n", border=0)

    pdf.output(pdf_path)

    print(f"\nReport salvati in: \n- {excel_path}\n- {pdf_path}")

if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument("--path", required=True, help="Cartella da scansionare")
    parser.add_argument("--max-size", type=int, default=10, help="Dimensione max file (MB)")
    parser.add_argument("--search", help="Parola o regex da cercare")
    parser.add_argument("--output", default=".", help="Cartella base per i report (verrà creata una sottocartella con data)")
    args = parser.parse_args()

    path_to_scan = args.path
    max_bytes = args.max_size * 1024 * 1024
    results = scan_directory(path_to_scan, max_bytes, args.search)
    if results:
        export_results(results, args.output)
    else:
        print("\nNessun dato analizzabile trovato.")

